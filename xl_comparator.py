# -*- coding: utf-8 -*-
"""
Created on Tue Mar  2 09:20:39 2021

@author: ashmanov
"""

# import pandas
import openpyxl
# from table_select import table_select
from tabulate import tabulate


def select_sheet(filename1, filename2):
    file1 = openpyxl.load_workbook(filename1, data_only=True)
    file2 = openpyxl.load_workbook(filename2, data_only=True)
    sheets_file1 = [sheet.title for sheet in file1.worksheets]
    sheets_file2 = [sheet.title for sheet in file2.worksheets]
    if any([sheets_file1 != sheets_file2,
            all([len(sheets_file1) == len(sheets_file2),
                 len(sheets_file1) > 1, len(sheets_file2) > 1])]):
        dict1 = {filename1: [str(i) + ". " + title for i, title in enumerate(sheets_file1, 1)],
                 filename2: [str(i) + ". " + title for i, title in enumerate(sheets_file2, 1)]}
        header1 = len_filename(filename1)
        header2 = len_filename(filename2)
        print(tabulate(dict1, headers=[header1, header2], tablefmt="pretty", colalign=("left", "left")), end="\n" * 2)
        print("Нужно выбрать, какие листы сравнивать:")
        sheet1 = select(1, file1)
        sheet2 = select(2, file2)
    else:
        sheet1 = file1.active
        sheet2 = file2.active
    return sheet1, sheet2


def len_filename(filename):
    if len(filename) > 15:
        return filename1[0:11] + " ... " + filename[-15:]
    else:
        return filename


def select(index, file):
    if index == 1:
        prnt = "Лист из первого файла: "
    elif index == 2:
        prnt = "Лист из второго файла: "
    while True:
        select = input(prnt)
        if select in [sheet.title for sheet in file.worksheets]:
            return file[select]
        elif select.isdigit():
            try:
                return file.worksheets[int(select) - 1]
            except IndexError:
                print("Такого листа в файле нет. Попробуйте ещё")
        else:
            print("Такого листа в файле нет. Попробуйте ещё")


def row_compare(row1, row2):
    compare = []
    for num_cell in range(min(len(row1), len(row2))):
        if row1[num_cell].value == row2[num_cell].value:
            compare.append(True)
        else:
            compare.append(False)
    if False in compare:
        return False
    else:
        return True


filename1 = "Волжский парк, котлован 3.2 от СКАЛА ООО (2).xlsx"
filename2 = "Волжский парк, котлован 3.2 от СКАЛА ООО (3).xlsx"
# filename1 = "file1.xlsx"
# filename2 = "file2.xlsx"
# result_file = openpyxl.Workbook()
# result_file.active.title = "Результат обработки"
# result_file.active.sheet_properties.tabColor = "FF0000"
sheet1, sheet2 = select_sheet(filename1, filename2)

equal_row = []
diff_row = []
for num_row_left in range(1, sheet1.max_row + 1):
    print(num_row_left)
    for num_row_right in range(1, sheet2.max_row + 1):
        if row_compare(sheet1[num_row_left], sheet2[num_row_right]):  # Строки одинаковы
            equal_row.append((num_row_left, num_row_right))
            break
        else:  # Строки не одинаковы
            pass

# num_row_left = 22
# num_row_right = 16
# if row_compare(sheet1[num_row_left], sheet2[num_row_right]):   # Строки одинаковы
#     equal_row.append((num_row_left, num_row_right))
# else:                                               # Строки не одинаковы
#     pass
