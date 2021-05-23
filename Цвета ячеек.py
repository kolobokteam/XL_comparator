# -*- coding: utf-8 -*-
"""
Created on Sun Mar  7 19:06:13 2021

@author: admOS
"""

import openpyxl #Подключаем библиотеку

wb = openpyxl.Workbook() #Создали книгу
work_sheet = wb.create_sheet(title='Test sheet') #Создали лист с названием и сделали его активным
a = int(input('Введите а: '))
cell_a = 'A'
i = 1
while i <= a:
    i = str(i)
    cell = cell_a + i
    work_sheet[cell] = 'Test text'
    work_sheet[cell].font = openpyxl.styles.Font(size=i)
    work_sheet[cell].fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='ff0000')
    i = int(i)
    i = i + 1
wb.save('ok.xlsx')